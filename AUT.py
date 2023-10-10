import openpyxl
from openpyxl.styles import NamedStyle

default_style = NamedStyle(name='default')
default_style.font.bold = True
default_style.fill.fill_type = "solid"
default_style.fill.start_color = "FFFF00"

workbook = openpyxl.Workbook()
workbook.default_style = default_style

arquivo_excel = openpyxl.load_workbook('/home/jovyan/aut/CCM.xlsx')

ccm = arquivo_excel['NFD X Avalara - Stefanini']

for linha in ccm.iter_rows(values_only=True):
    print(linha)
    
